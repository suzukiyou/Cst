import ezdxf
import openpyxl

if __name__ == "__main__":
    filename = ".dxf"
    xlsxfilename = ".xlsx"
    dwg = ezdxf.readfile(filename)
    xMin = 3559
    yMin = 9808
    xMax = 80510
    yMax = 55802

    def xisInnerRegion(x): return x > xMin and x < xMax

    def yisInnerRegion(x): return x > yMin and x < yMax

    def isInnerLine(ent):
        return xisInnerRegion(ent.dxf.start[0]) and xisInnerRegion(ent.dxf.end[0]) \
            and yisInnerRegion(ent.dxf.start[1]) and yisInnerRegion(ent.dxf.end[1])
    lines = [l for l in dwg.query("LINE") if isInnerLine(l)]

    def isHorizontal(l):
        return abs(l.dxf.start[1]-l.dxf.end[1]) < 1

    def isVertical(l):
        return abs(l.dxf.start[0]-l.dxf.end[0]) < 1

    hlines = sorted(set([l.dxf.start[1]
                         for l in lines if isHorizontal(l)]), reverse=True)
    vlines = sorted(set([l.dxf.start[0]
                         for l in lines if isVertical(l)]))

    rows = len(hlines)-1
    cols = len(vlines)-1

    def hindex(y):
        for i, xl in enumerate(hlines):
            if xl < y:
                return i

    def vindex(x):
        for i, yl in enumerate(vlines):
            if yl > x:
                return i

    def isInnerText(ent):
        return xisInnerRegion(ent.dxf.insert[0]) and yisInnerRegion(ent.dxf.insert[1])

    textmap = {}
    for tent in dwg.query("TEXT"):
        if isInnerText(tent):
            textmap[(hindex(tent.dxf.insert[1]), vindex(
                tent.dxf.insert[0]))] = tent.dxf.text

    wb = openpyxl.Workbook()
    ws = wb.active
    for key, text in textmap.items():
        if ws.cell(key[0], key[1]).value == None:
            ws.cell(key[0], key[1]).value = text
        else:
            ws.cell(key[0], key[1]).value = ws.cell(key[0], key[1]).value+text
    wb.save(xlsxfilename)
